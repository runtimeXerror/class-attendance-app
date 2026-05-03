from fastapi import FastAPI, Depends, HTTPException, Response, UploadFile, File, Form, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import date, datetime, timezone, timedelta
from typing import List, Optional
from io import BytesIO

from . import models, schemas, auth
from .database import engine, get_db, Base

# Render servers run in UTC. The school is in India, so all human-facing
# timestamps (Excel/PDF "Generated" footer, edited-at, etc.) need IST.
IST = timezone(timedelta(hours=5, minutes=30))

def now_ist():
    return datetime.now(IST)

Base.metadata.create_all(bind=engine)


# ------------------------------------------------------------------
# Lightweight startup migration: backfill teacher_branches so that
# every existing teacher is linked to their primary branch_id.
# Postgres `create_all` adds the new table, but pre-existing rows
# need this one-time backfill to be visible in HOD listings.
# ------------------------------------------------------------------
def _ensure_teacher_branch_backfill():
    from .database import SessionLocal
    db = SessionLocal()
    try:
        rows = db.query(models.Teacher).all()
        existing = {
            (tb.teacher_id, tb.branch_id)
            for tb in db.query(models.TeacherBranch).all()
        }
        created = 0
        for t in rows:
            if t.branch_id and (t.id, t.branch_id) not in existing:
                db.add(models.TeacherBranch(teacher_id=t.id, branch_id=t.branch_id))
                created += 1
        if created:
            db.commit()
    except Exception:
        db.rollback()
    finally:
        db.close()


_ensure_teacher_branch_backfill()


app = FastAPI(title="Class Attendance System API v1.2")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.api_route("/", methods=["GET", "HEAD"])
def root():
    return {"status": "ok", "message": "Class Attendance API v1.1 is running"}


# Ultra-fast keep-alive endpoint — no DB, no auth, no body work.
# Hit this from the mobile app every few minutes to stop Render's
# free plan from sleeping the dyno (15 min idle -> 30-50s cold start).
# Accepts HEAD too — UptimeRobot's default HTTP(s) check uses HEAD,
# without this it gets a 405 and reports the service as Down.
@app.api_route("/healthz", methods=["GET", "HEAD"])
def healthz():
    return {"ok": True}


# ============ LOGIN ============
@app.post("/api/login", response_model=schemas.TokenResponse)
def login(req: schemas.LoginRequest, db: Session = Depends(get_db)):
    if req.role == "superadmin":
        if not req.email or not req.password:
            raise HTTPException(400, "Email and password required")
        sa = db.query(models.SuperAdmin).filter(models.SuperAdmin.email == req.email).first()
        if not sa or not auth.verify_password(req.password, sa.password_hash):
            raise HTTPException(401, "Invalid credentials")
        token = auth.create_token({"id": sa.id, "role": "superadmin"})
        return {"access_token": token, "role": "superadmin", "name": sa.name, "user_id": sa.id}

    if req.role == "admin":
        if not req.email or not req.password:
            raise HTTPException(400, "Email and password required")
        admin = db.query(models.Admin).filter(models.Admin.email == req.email).first()
        if not admin or not auth.verify_password(req.password, admin.password_hash):
            raise HTTPException(401, "Invalid credentials")
        if not admin.is_active:
            raise HTTPException(403, "Account deactivated")
        token = auth.create_token({"id": admin.id, "role": "admin"})
        branch = db.query(models.Branch).filter(models.Branch.id == admin.branch_id).first()
        return {
            "access_token": token, "role": "admin", "name": admin.name, "user_id": admin.id,
            "branch_name": branch.name if branch else None,
            "branch_code": branch.code if branch else None,
            "must_change_password": not admin.password_changed,
        }

    if req.role == "teacher":
        if not req.email or not req.password:
            raise HTTPException(400, "Email and password required")
        teacher = db.query(models.Teacher).filter(models.Teacher.email == req.email).first()
        if not teacher or not auth.verify_password(req.password, teacher.password_hash):
            raise HTTPException(401, "Invalid credentials")
        if not teacher.is_active:
            raise HTTPException(403, "Account deactivated")
        token = auth.create_token({"id": teacher.id, "role": "teacher"})
        branch = db.query(models.Branch).filter(models.Branch.id == teacher.branch_id).first()
        return {
            "access_token": token, "role": "teacher", "name": teacher.name, "user_id": teacher.id,
            "branch_name": branch.name if branch else None,
            "branch_code": branch.code if branch else None,
            "must_change_password": not teacher.password_changed,
            "profile_image": teacher.profile_image,
        }

    if req.role == "student":
        if not req.reg_no or not req.dob:
            raise HTTPException(400, "Reg No and DOB required")
        student = db.query(models.Student).filter(
            models.Student.reg_no == req.reg_no,
            models.Student.date_of_birth == req.dob
        ).first()
        if not student:
            raise HTTPException(401, "Invalid Reg No or DOB")
        token = auth.create_token({"id": student.id, "role": "student"})
        branch = db.query(models.Branch).filter(models.Branch.id == student.branch_id).first()
        return {
            "access_token": token, "role": "student", "name": student.name, "user_id": student.id,
            "branch_name": branch.name if branch else None,
            "branch_code": branch.code if branch else None,
        }

    raise HTTPException(400, "Invalid role")


# ============ PASSWORD CHANGE ============
@app.post("/api/change-password")
def change_password(
    data: schemas.PasswordChange,
    current=Depends(auth.get_current_user),
    db: Session = Depends(get_db)
):
    user = current["user"]
    role = current["role"]
    if not auth.verify_password(data.old_password, user.password_hash):
        raise HTTPException(401, "Old password incorrect")
    user.password_hash = auth.hash_password(data.new_password)
    if role in ["admin", "teacher"]:
        user.password_changed = True
    db.commit()
    return {"success": True}


# ============ PUBLIC ============
@app.get("/api/batches", response_model=List[schemas.BatchOut])
def list_batches(db: Session = Depends(get_db)):
    return db.query(models.Batch).filter(models.Batch.is_active == True).order_by(models.Batch.start_year.desc()).all()


@app.get("/api/branches", response_model=List[schemas.BranchOut])
def list_branches(db: Session = Depends(get_db)):
    return db.query(models.Branch).all()


# ============ SUPER ADMIN ============
@app.post("/api/superadmin/admins", response_model=schemas.AdminCreateResponse)
def create_admin_by_super(
    data: schemas.AdminCreate,
    background_tasks: BackgroundTasks,
    current=Depends(auth.require_role("superadmin")),
    db: Session = Depends(get_db)
):
    # Auto-generate default password (can be overridden)
    import secrets, string
    if data.password:
        default_pwd = data.password
    else:
        alphabet = string.ascii_letters + string.digits
        default_pwd = 'A' + ''.join(secrets.choice(alphabet) for _ in range(7))

    admin = models.Admin(
        email=data.email,
        password_hash=auth.hash_password(default_pwd),
        name=data.name,
        branch_id=data.branch_id,
        phone=data.phone,
        password_changed=False,   # force change on first login
    )
    db.add(admin)
    try:
        db.commit()
        db.refresh(admin)
    except Exception:
        db.rollback()
        raise HTTPException(400, "Email already exists")

    branch = db.query(models.Branch).filter(models.Branch.id == admin.branch_id).first()

    # Email send is queued as a background task so the API responds in
    # ~50 ms even if Gmail SMTP takes 20+ s. The super-admin's screen never
    # spins waiting on a network round-trip we don't control.
    from . import email_utils
    smtp_set = email_utils.is_configured()
    if smtp_set:
        background_tasks.add_task(
            email_utils.send_credentials_email,
            admin.email, admin.name, "Branch HOD", default_pwd,
        )

    return {
        "id": admin.id,
        "email": admin.email,
        "name": admin.name,
        "branch_code": branch.code if branch else "",
        # If SMTP is configured we trust it'll go through (logs will show
        # success/failure). Hide the password from the screen in that case.
        # If SMTP isn't configured at all we surface the password so it can
        # be shared manually.
        "default_password": "" if smtp_set else default_pwd,
        "email_sent": smtp_set,
        "message": (
            f"HOD created. Credentials emailing to {admin.email} (check inbox in ~10 s)."
            if smtp_set
            else f"HOD created. Default password: {default_pwd}\n"
                 f"⚠ SMTP is not configured — share this password manually."
        ),
    }


@app.get("/api/superadmin/admins", response_model=List[schemas.AdminOut])
def list_all_admins(current=Depends(auth.require_role("superadmin")), db: Session = Depends(get_db)):
    branch_map = {b.id: b for b in db.query(models.Branch).all()}
    admins = db.query(models.Admin).order_by(models.Admin.branch_id).all()
    out = []
    for a in admins:
        d = schemas.AdminOut.from_orm(a)
        b = branch_map.get(a.branch_id)
        d.branch_name = b.name if b else None
        d.branch_code = b.code if b else None
        out.append(d)
    return out


@app.patch("/api/superadmin/admins/{admin_id}", response_model=schemas.AdminOut)
def update_admin(
    admin_id: int,
    data: schemas.AdminUpdate,
    current=Depends(auth.require_role("superadmin")),
    db: Session = Depends(get_db)
):
    admin = db.query(models.Admin).filter(models.Admin.id == admin_id).first()
    if not admin:
        raise HTTPException(404, "Admin not found")
    if data.name is not None: admin.name = data.name
    if data.phone is not None: admin.phone = data.phone
    if data.branch_id is not None: admin.branch_id = data.branch_id
    if data.is_active is not None: admin.is_active = data.is_active
    if data.password: admin.password_hash = auth.hash_password(data.password)
    db.commit()
    db.refresh(admin)
    return _enrich_admin(admin, db)


@app.delete("/api/superadmin/admins/{admin_id}")
def delete_admin(
    admin_id: int,
    current=Depends(auth.require_role("superadmin")),
    db: Session = Depends(get_db)
):
    admin = db.query(models.Admin).filter(models.Admin.id == admin_id).first()
    if not admin:
        raise HTTPException(404, "Admin not found")
    db.delete(admin)
    db.commit()
    return {"success": True}


def _enrich_admin(admin, db):
    branch = db.query(models.Branch).filter(models.Branch.id == admin.branch_id).first()
    d = schemas.AdminOut.from_orm(admin)
    d.branch_name = branch.name if branch else None
    d.branch_code = branch.code if branch else None
    return d


@app.post("/api/superadmin/batches", response_model=schemas.BatchOut)
def create_batch(
    data: schemas.BatchCreate,
    current=Depends(auth.require_role("superadmin")),
    db: Session = Depends(get_db)
):
    name = f"{data.start_year}-{str(data.end_year)[-2:]}"
    existing = db.query(models.Batch).filter(models.Batch.name == name).first()
    if existing:
        return existing
    b = models.Batch(name=name, start_year=data.start_year, end_year=data.end_year, current_semester=data.current_semester or 1)
    db.add(b)
    db.commit()
    db.refresh(b)
    return b


@app.get("/api/superadmin/stats")
def super_stats(current=Depends(auth.require_role("superadmin")), db: Session = Depends(get_db)):
    return {
        "total_branches": db.query(models.Branch).count(),
        "total_admins": db.query(models.Admin).count(),
        "total_teachers": db.query(models.Teacher).count(),
        "total_students": db.query(models.Student).count(),
        "total_subjects": db.query(models.Subject).count(),
        "total_batches": db.query(models.Batch).count(),
    }


# ============ ADMIN ============
def _teacher_in_admin_branch(db: Session, teacher_id: int, branch_id: int):
    """Helper: returns the Teacher if linked to this branch (via TeacherBranch
    or as primary branch_id), else None. Used by HOD endpoints so any HOD can
    edit/unlink a teacher that teaches in their branch — even one whose primary
    branch is elsewhere."""
    link = db.query(models.TeacherBranch).filter(
        models.TeacherBranch.teacher_id == teacher_id,
        models.TeacherBranch.branch_id == branch_id,
    ).first()
    if link:
        return db.query(models.Teacher).filter(models.Teacher.id == teacher_id).first()
    # Fallback for legacy rows (pre-migration) where TeacherBranch may be empty.
    return db.query(models.Teacher).filter(
        models.Teacher.id == teacher_id,
        models.Teacher.branch_id == branch_id,
    ).first()


@app.post("/api/admin/teachers", response_model=schemas.TeacherCreateResponse)
def create_teacher(
    data: schemas.TeacherCreate,
    background_tasks: BackgroundTasks,
    current=Depends(auth.require_role("admin")),
    db: Session = Depends(get_db)
):
    """Create a teacher OR link an existing teacher (by email) to this HOD's
    branch. Same teacher can teach in CSE + CSE-DS (or any other combo) and
    only receives the credentials email once — on FIRST creation."""
    admin = current["user"]

    existing = db.query(models.Teacher).filter(models.Teacher.email == data.email).first()
    if existing:
        # Already in this HOD's branch?
        link = db.query(models.TeacherBranch).filter(
            models.TeacherBranch.teacher_id == existing.id,
            models.TeacherBranch.branch_id == admin.branch_id,
        ).first()
        if link:
            raise HTTPException(400, "Teacher already exists in your branch")
        # Different branch — link them. No new password, no new email.
        db.add(models.TeacherBranch(teacher_id=existing.id, branch_id=admin.branch_id))
        # Optionally update phone if HOD provided one and existing has none
        if data.phone and not existing.phone:
            existing.phone = data.phone
        db.commit()
        return {
            "id": existing.id,
            "email": existing.email,
            "name": existing.name,
            "default_password": "",
            "email_sent": False,
            "message": (
                f"{existing.name} was already registered in another branch. "
                f"Linked to your branch — no new email/credentials sent. "
                f"They can use their existing login."
            ),
        }

    import secrets, string
    alphabet = string.ascii_letters + string.digits
    default_pwd = 'T' + ''.join(secrets.choice(alphabet) for _ in range(7))
    teacher = models.Teacher(
        email=data.email,
        password_hash=auth.hash_password(default_pwd),
        name=data.name,
        phone=data.phone,
        branch_id=admin.branch_id,
        created_by_admin_id=admin.id,
        password_changed=False,
    )
    db.add(teacher)
    try:
        db.commit()
        db.refresh(teacher)
    except Exception:
        db.rollback()
        raise HTTPException(400, "Email already exists")

    # Mirror to teacher_branches for the brand-new teacher
    db.add(models.TeacherBranch(teacher_id=teacher.id, branch_id=admin.branch_id))
    db.commit()

    # Email send goes to a background task — API responds in milliseconds
    # even if Gmail SMTP takes 20+ s (or fails entirely). HOD ko UI
    # blocking 40-second wait nahi milta.
    from . import email_utils
    smtp_set = email_utils.is_configured()
    if smtp_set:
        background_tasks.add_task(
            email_utils.send_credentials_email,
            teacher.email, teacher.name, "Teacher", default_pwd,
        )

    return {
        "id": teacher.id,
        "email": teacher.email,
        "name": teacher.name,
        # Hide password from HOD's screen when SMTP is configured (email
        # will deliver). Surface it only when SMTP isn't set, so HOD has
        # something to share manually.
        "default_password": "" if smtp_set else default_pwd,
        "email_sent": smtp_set,
        "message": (
            f"Teacher created. Credentials emailing to {teacher.email} (check inbox in ~10 s)."
            if smtp_set
            else f"Teacher created. Default password: {default_pwd}\n"
                 f"⚠ SMTP is not configured — share this password manually."
        ),
    }


@app.get("/api/admin/teachers", response_model=List[schemas.TeacherOut])
def list_teachers(current=Depends(auth.require_role("admin")), db: Session = Depends(get_db)):
    """All teachers attached to this HOD's branch — including cross-branch
    teachers whose primary branch is different but who teach a subject here."""
    admin = current["user"]
    # Eager bulk fetch: one query for all teacher ids in this branch, then one
    # for the teacher rows. No N+1.
    linked_ids = [
        tb.teacher_id for tb in
        db.query(models.TeacherBranch).filter(
            models.TeacherBranch.branch_id == admin.branch_id
        ).all()
    ]
    if not linked_ids:
        return []
    return db.query(models.Teacher).filter(models.Teacher.id.in_(linked_ids)).all()


@app.patch("/api/admin/teachers/{teacher_id}", response_model=schemas.TeacherOut)
def update_teacher(
    teacher_id: int,
    data: schemas.TeacherUpdate,
    current=Depends(auth.require_role("admin")),
    db: Session = Depends(get_db)
):
    admin = current["user"]
    teacher = _teacher_in_admin_branch(db, teacher_id, admin.branch_id)
    if not teacher:
        raise HTTPException(404, "Teacher not found")
    if data.name is not None: teacher.name = data.name
    if data.phone is not None: teacher.phone = data.phone
    if data.is_active is not None: teacher.is_active = data.is_active
    if data.profile_image is not None: teacher.profile_image = data.profile_image
    db.commit()
    db.refresh(teacher)
    return teacher


@app.delete("/api/admin/teachers/{teacher_id}")
def delete_teacher(
    teacher_id: int,
    current=Depends(auth.require_role("admin")),
    db: Session = Depends(get_db)
):
    """Unlink teacher from THIS HOD's branch. If the teacher still teaches
    in another branch (TeacherBranch row exists) the user account stays
    intact — only the link is removed. Otherwise the teacher is fully deleted."""
    admin = current["user"]
    teacher = _teacher_in_admin_branch(db, teacher_id, admin.branch_id)
    if not teacher:
        raise HTTPException(404, "Teacher not found")

    # Block delete if the teacher still has subjects in this branch.
    has_subjects_here = db.query(models.Subject).filter(
        models.Subject.teacher_id == teacher_id,
        models.Subject.branch_id == admin.branch_id,
    ).first()
    if has_subjects_here:
        raise HTTPException(
            400,
            "Cannot remove: teacher is assigned to subjects in your branch. "
            "Reassign or delete those subjects first."
        )

    # Drop the link for this branch
    db.query(models.TeacherBranch).filter(
        models.TeacherBranch.teacher_id == teacher_id,
        models.TeacherBranch.branch_id == admin.branch_id,
    ).delete()

    # Any other branches still using this teacher?
    other_links = db.query(models.TeacherBranch).filter(
        models.TeacherBranch.teacher_id == teacher_id,
    ).count()

    if other_links > 0:
        # Keep the teacher account; if their primary branch was THIS branch,
        # promote one of the remaining branches to primary so logins still work.
        if teacher.branch_id == admin.branch_id:
            new_primary = db.query(models.TeacherBranch).filter(
                models.TeacherBranch.teacher_id == teacher_id
            ).first()
            teacher.branch_id = new_primary.branch_id
        db.commit()
        return {"success": True, "unlinked_only": True}

    # No links left → fully delete (cascade attendance/subjects-elsewhere cleanup)
    db.query(models.AttendanceRecord).filter(
        models.AttendanceRecord.marked_by_teacher_id == teacher_id
    ).update({models.AttendanceRecord.marked_by_teacher_id: None})
    db.delete(teacher)
    db.commit()
    return {"success": True}


# ---- Bulk student import (HOD uploads CSV/XLSX from Google Form export) ----
class BulkStudentRow(BaseModel):
    reg_no: str
    name: str
    date_of_birth: date
    student_type: Optional[str] = "regular"
    phone: Optional[str] = None
    current_semester: Optional[int] = None  # falls back to body.current_semester


class BulkImportRequest(BaseModel):
    batch_id: int
    current_semester: int
    rows: List[BulkStudentRow]


# ---- File-based bulk import: HOD uploads a .csv / .xlsx straight from
#      Google Form export. Single endpoint handles both formats so the
#      mobile UI just hands over the picked file and forgets about parsing. ----

def _norm_header(s: str) -> str:
    return "".join(ch for ch in str(s or "").lower() if ch.isalnum())


_HEADER_ALIASES = {
    "reg_no":       {"regno", "registration", "registrationno", "registrationnumber",
                     "enrollmentno", "enrollment", "rollno", "roll"},
    "name":         {"name", "fullname", "studentname"},
    "dob":          {"dob", "dateofbirth", "birthdate", "birthday"},
    "type":         {"type", "studenttype", "category"},
    "phone":        {"phone", "mobile", "contact", "phoneno", "mobileno", "whatsapp"},
    "semester":     {"semester", "currentsemester", "sem"},
}


def _detect_columns(headers: List[str]) -> dict:
    norm = [_norm_header(h) for h in headers]
    out = {}
    for key, aliases in _HEADER_ALIASES.items():
        out[key] = next((i for i, h in enumerate(norm) if h in aliases), -1)
    return out


def _parse_dob(s) -> Optional[date]:
    """Liberal DOB parser. Sheets/Forms emit DD/MM/YYYY most often, but
    we also accept ISO and a few other separators. Returns None on fail."""
    if s is None:
        return None
    if isinstance(s, date) and not isinstance(s, datetime):
        return s
    if isinstance(s, datetime):
        return s.date()
    s = str(s).strip()
    if not s:
        return None
    import re
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    m = re.match(r"^(\d{1,2})[\/\-\.](\d{1,2})[\/\-\.](\d{2,4})", s)
    if m:
        d, mo, y = m.group(1), m.group(2), m.group(3)
        if len(y) == 2:
            y = "20" + y
        try:
            return date(int(y), int(mo), int(d))
        except ValueError:
            return None
    return None


def _normalize_type(v) -> str:
    s = _norm_header(v)
    if not s:
        return "regular"
    if "lateral" in s or s == "le":
        return "lateral_entry"
    if "back" in s:
        return "back_year"
    return "regular"


@app.post("/api/admin/students/bulk-upload")
async def bulk_upload_students(
    batch_id: int = Form(...),
    current_semester: int = Form(...),
    file: UploadFile = File(...),
    current=Depends(auth.require_role("admin")),
    db: Session = Depends(get_db)
):
    """HOD picks a CSV / XLSX file (e.g. Google Sheets → Download → CSV).
    Required header columns: reg_no, name, dob. Optional: type, phone, semester.
    Header names are matched case-insensitively with a small alias list."""
    admin = current["user"]
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "Empty file")

    fname = (file.filename or "").lower()
    rows_data: List[List] = []
    try:
        if fname.endswith(".xlsx") or fname.endswith(".xls"):
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(raw), data_only=True, read_only=True)
            ws = wb.active
            for r in ws.iter_rows(values_only=True):
                if any(cell is not None and str(cell).strip() != "" for cell in r):
                    rows_data.append(list(r))
        else:
            # CSV / TSV — sniff delimiter on header line
            import csv as _csv, io as _io
            text = raw.decode("utf-8-sig", errors="replace")
            first_line = text.split("\n", 1)[0]
            delim = "\t" if "\t" in first_line and "," not in first_line else ","
            reader = _csv.reader(_io.StringIO(text), delimiter=delim)
            for row in reader:
                if any((c or "").strip() for c in row):
                    rows_data.append(row)
    except Exception as e:
        raise HTTPException(400, f"Could not parse file: {e}")

    if len(rows_data) < 2:
        raise HTTPException(400, "Need a header row + at least one data row.")

    cols = _detect_columns(rows_data[0])
    if cols["reg_no"] == -1 or cols["name"] == -1 or cols["dob"] == -1:
        raise HTTPException(400, "Missing required columns. Need reg_no, name, dob.")

    # Reuse the same insertion logic as the JSON endpoint via a virtual list
    parsed: List[BulkStudentRow] = []
    parse_warnings: List[dict] = []
    for li, row in enumerate(rows_data[1:], start=2):
        def cell(i):
            return row[i] if i != -1 and i < len(row) else None
        reg = (str(cell(cols["reg_no"]) or "").strip())
        nm = (str(cell(cols["name"]) or "").strip())
        dob_val = _parse_dob(cell(cols["dob"]))
        if not reg or not nm or not dob_val:
            parse_warnings.append({"row": li, "reason": "Missing reg_no / name / dob"})
            continue
        sem_raw = cell(cols["semester"])
        try:
            sem = int(sem_raw) if sem_raw is not None and str(sem_raw).strip() else None
        except (TypeError, ValueError):
            sem = None
        parsed.append(BulkStudentRow(
            reg_no=reg,
            name=nm,
            date_of_birth=dob_val,
            student_type=_normalize_type(cell(cols["type"])),
            phone=(str(cell(cols["phone"]) or "").strip() or None),
            current_semester=sem,
        ))

    if not parsed:
        return {
            "added": 0, "skipped_existing": 0,
            "errors": parse_warnings, "total_rows": len(rows_data) - 1,
            "parse_warnings": parse_warnings,
        }

    # Delegate to the same insertion logic
    body = BulkImportRequest(
        batch_id=batch_id, current_semester=current_semester, rows=parsed
    )
    result = bulk_import_students(body, current, db)
    if parse_warnings:
        result["parse_warnings"] = parse_warnings
    return result


@app.post("/api/admin/students/bulk")
def bulk_import_students(
    body: BulkImportRequest,
    current=Depends(auth.require_role("admin")),
    db: Session = Depends(get_db)
):
    """HOD bulk-uploads students collected from a Google Form / spreadsheet.
    Existing reg_nos are skipped (not failed) so the same upload can be
    re-run safely after corrections. Auto-enrolls in current semester subjects."""
    admin = current["user"]

    # Pre-load existing reg_nos to skip dup rows fast
    existing_regs = {
        r[0] for r in
        db.query(models.Student.reg_no).filter(
            models.Student.reg_no.in_([r.reg_no for r in body.rows])
        ).all()
    }

    subjects = db.query(models.Subject).filter(
        models.Subject.branch_id == admin.branch_id,
        models.Subject.batch_id == body.batch_id,
        models.Subject.semester == body.current_semester,
    ).all()

    added, skipped_existing, errors = 0, 0, []
    for idx, row in enumerate(body.rows):
        try:
            if not row.reg_no or not row.name:
                errors.append({"row": idx + 1, "reason": "Missing reg_no or name"})
                continue
            if row.reg_no in existing_regs:
                skipped_existing += 1
                continue
            stype = row.student_type or "regular"
            if stype not in ("regular", "lateral_entry", "back_year"):
                stype = "regular"
            sem = row.current_semester or body.current_semester
            student = models.Student(
                reg_no=row.reg_no.strip(),
                name=row.name.strip(),
                date_of_birth=row.date_of_birth,
                branch_id=admin.branch_id,
                batch_id=body.batch_id,
                current_semester=sem,
                student_type=stype,
                phone=(row.phone or "").strip() or None,
            )
            db.add(student)
            db.flush()
            existing_regs.add(row.reg_no)
            for subj in subjects:
                if subj.semester == sem:
                    db.add(models.Enrollment(student_id=student.id, subject_id=subj.id))
            added += 1
        except Exception as e:
            errors.append({"row": idx + 1, "reason": str(e)[:120]})
            db.rollback()

    db.commit()
    return {
        "added": added,
        "skipped_existing": skipped_existing,
        "errors": errors,
        "total_rows": len(body.rows),
    }


@app.post("/api/admin/students", response_model=schemas.StudentOut)
def create_student(
    data: schemas.StudentCreate,
    current=Depends(auth.require_role("admin")),
    db: Session = Depends(get_db)
):
    admin = current["user"]
    student = models.Student(
        reg_no=data.reg_no,
        name=data.name,
        date_of_birth=data.date_of_birth,
        branch_id=admin.branch_id,
        batch_id=data.batch_id,
        current_semester=data.current_semester,
        student_type=data.student_type,
        phone=data.phone,
    )
    db.add(student)
    try:
        db.commit()
        db.refresh(student)
    except Exception:
        db.rollback()
        raise HTTPException(400, "Reg No already exists")

    subjects = db.query(models.Subject).filter(
        models.Subject.branch_id == admin.branch_id,
        models.Subject.batch_id == data.batch_id,
        models.Subject.semester == data.current_semester,
    ).all()
    for subj in subjects:
        db.add(models.Enrollment(student_id=student.id, subject_id=subj.id))
    db.commit()
    return student


@app.get("/api/admin/students", response_model=List[schemas.StudentOut])
def list_students(
    batch_id: Optional[int] = None,
    semester: Optional[int] = None,
    current=Depends(auth.require_role("admin")),
    db: Session = Depends(get_db)
):
    admin = current["user"]
    q = db.query(models.Student).filter(models.Student.branch_id == admin.branch_id)
    if batch_id:
        q = q.filter(models.Student.batch_id == batch_id)
    if semester:
        q = q.filter(models.Student.current_semester == semester)
    return q.order_by(models.Student.reg_no).all()


@app.patch("/api/admin/students/{student_id}", response_model=schemas.StudentOut)
def update_student(
    student_id: int,
    data: schemas.StudentUpdate,
    current=Depends(auth.require_role("admin")),
    db: Session = Depends(get_db)
):
    admin = current["user"]
    student = db.query(models.Student).filter(
        models.Student.id == student_id,
        models.Student.branch_id == admin.branch_id
    ).first()
    if not student:
        raise HTTPException(404, "Student not found")
    # Allow reg_no change but ensure uniqueness
    if data.reg_no is not None and data.reg_no != student.reg_no:
        clash = db.query(models.Student).filter(
            models.Student.reg_no == data.reg_no,
            models.Student.id != student_id
        ).first()
        if clash:
            raise HTTPException(400, f"Registration number {data.reg_no} already exists")
        student.reg_no = data.reg_no
    if data.name is not None: student.name = data.name
    if data.date_of_birth is not None: student.date_of_birth = data.date_of_birth
    if data.current_semester is not None: student.current_semester = data.current_semester
    if data.student_type is not None: student.student_type = data.student_type
    if data.phone is not None: student.phone = data.phone
    if data.is_active is not None: student.is_active = data.is_active
    db.commit()
    db.refresh(student)
    return student


@app.delete("/api/admin/students/{student_id}")
def delete_student(
    student_id: int,
    current=Depends(auth.require_role("admin")),
    db: Session = Depends(get_db)
):
    admin = current["user"]
    student = db.query(models.Student).filter(
        models.Student.id == student_id,
        models.Student.branch_id == admin.branch_id
    ).first()
    if not student:
        raise HTTPException(404, "Student not found")
    db.query(models.Enrollment).filter(models.Enrollment.student_id == student_id).delete()
    db.query(models.AttendanceRecord).filter(models.AttendanceRecord.student_id == student_id).delete()
    db.delete(student)
    db.commit()
    return {"success": True}


@app.post("/api/admin/subjects", response_model=schemas.SubjectOut)
def create_subject(
    data: schemas.SubjectCreate,
    current=Depends(auth.require_role("admin")),
    db: Session = Depends(get_db)
):
    admin = current["user"]

    # Auto-link the assigned teacher to this branch if they aren't already.
    # Lets a HOD assign a teacher whose primary branch is elsewhere — common
    # for shared faculty (e.g. Sankalp Sonu teaching CSE + CSE-DS).
    teacher = db.query(models.Teacher).filter(models.Teacher.id == data.teacher_id).first()
    if not teacher:
        raise HTTPException(404, "Teacher not found")
    has_link = db.query(models.TeacherBranch).filter(
        models.TeacherBranch.teacher_id == teacher.id,
        models.TeacherBranch.branch_id == admin.branch_id,
    ).first()
    if not has_link:
        db.add(models.TeacherBranch(teacher_id=teacher.id, branch_id=admin.branch_id))
        db.commit()

    subject = models.Subject(
        code=data.code,
        name=data.name,
        branch_id=admin.branch_id,
        batch_id=data.batch_id,
        semester=data.semester,
        teacher_id=data.teacher_id,
        credits=data.credits or 3,
        session=data.session,
    )
    db.add(subject)
    db.commit()
    db.refresh(subject)

    students = db.query(models.Student).filter(
        models.Student.branch_id == admin.branch_id,
        models.Student.batch_id == data.batch_id,
        models.Student.current_semester == data.semester,
        models.Student.is_active == True
    ).all()
    for s in students:
        db.add(models.Enrollment(student_id=s.id, subject_id=subject.id))
    db.commit()
    return _enrich_subject(subject, db)


@app.get("/api/admin/subjects", response_model=List[schemas.SubjectOut])
def list_subjects(
    batch_id: Optional[int] = None,
    current=Depends(auth.require_role("admin")),
    db: Session = Depends(get_db)
):
    admin = current["user"]
    q = db.query(models.Subject).filter(models.Subject.branch_id == admin.branch_id)
    if batch_id:
        q = q.filter(models.Subject.batch_id == batch_id)
    return [_enrich_subject(s, db) for s in q.all()]


@app.delete("/api/admin/subjects/{subject_id}")
def delete_subject(
    subject_id: int,
    current=Depends(auth.require_role("admin")),
    db: Session = Depends(get_db)
):
    admin = current["user"]
    subject = db.query(models.Subject).filter(
        models.Subject.id == subject_id,
        models.Subject.branch_id == admin.branch_id
    ).first()
    if not subject:
        raise HTTPException(404, "Subject not found")
    db.query(models.Enrollment).filter(models.Enrollment.subject_id == subject_id).delete()
    db.query(models.AttendanceRecord).filter(models.AttendanceRecord.subject_id == subject_id).delete()
    db.delete(subject)
    db.commit()
    return {"success": True}


def _enrich_subject(subject, db):
    teacher = db.query(models.Teacher).filter(models.Teacher.id == subject.teacher_id).first()
    batch = db.query(models.Batch).filter(models.Batch.id == subject.batch_id).first()
    o = schemas.SubjectOut.from_orm(subject)
    o.teacher_name = teacher.name if teacher else None
    o.batch_name = batch.name if batch else None
    return o


# ============ TEACHER ============
@app.get("/api/teacher/subjects", response_model=List[schemas.SubjectOut])
def teacher_subjects(current=Depends(auth.require_role("teacher")), db: Session = Depends(get_db)):
    teacher = current["user"]
    subjects = db.query(models.Subject).filter(models.Subject.teacher_id == teacher.id).all()
    return [_enrich_subject(s, db) for s in subjects]


@app.get("/api/teacher/subjects/{subject_id}/students")
def students_in_subject(
    subject_id: int,
    current=Depends(auth.require_role("teacher")),
    db: Session = Depends(get_db)
):
    teacher = current["user"]
    subj = db.query(models.Subject).filter(
        models.Subject.id == subject_id,
        models.Subject.teacher_id == teacher.id
    ).first()
    if not subj:
        raise HTTPException(404, "Subject not found")

    enrollments = db.query(models.Enrollment).filter(models.Enrollment.subject_id == subject_id).all()
    student_ids = [e.student_id for e in enrollments]
    students = db.query(models.Student).filter(
        models.Student.id.in_(student_ids),
        models.Student.is_active == True
    ).all()

    # Sort: Back Year → Regular → Lateral Entry, then by reg_no within each
    type_order = {"back_year": 0, "regular": 1, "lateral_entry": 2}
    students.sort(key=lambda s: (type_order.get(s.student_type, 99), s.reg_no or ""))

    return [{
        "id": s.id, "reg_no": s.reg_no, "name": s.name,
        "student_type": s.student_type, "is_verified": s.is_verified
    } for s in students]


@app.post("/api/teacher/attendance/mark")
def mark_attendance(
    data: schemas.AttendanceMark,
    current=Depends(auth.require_role("teacher")),
    db: Session = Depends(get_db)
):
    teacher = current["user"]
    subj = db.query(models.Subject).filter(
        models.Subject.id == data.subject_id,
        models.Subject.teacher_id == teacher.id
    ).first()
    if not subj:
        raise HTTPException(404, "Subject not assigned to you")

    db.query(models.AttendanceRecord).filter(
        models.AttendanceRecord.subject_id == data.subject_id,
        models.AttendanceRecord.class_date == data.class_date
    ).delete()

    # If every student is marked Absent, treat the date as "no class held"
    # — drop the records entirely so the day disappears from the calendar
    # and Excel/PDF exports. Lets teachers undo a wrongly-marked day by
    # selecting it, hitting "All Absent", and submitting.
    has_present = any(m.get("status") == "P" for m in data.marks)
    if not has_present:
        db.commit()
        return {"success": True, "count": 0, "deleted_day": True}

    for m in data.marks:
        rec = models.AttendanceRecord(
            student_id=m["student_id"],
            subject_id=data.subject_id,
            class_date=data.class_date,
            status=m["status"],
            marked_by_teacher_id=teacher.id,
        )
        db.add(rec)
    db.commit()
    return {"success": True, "count": len(data.marks)}


@app.get("/api/teacher/attendance/{subject_id}")
def get_attendance(
    subject_id: int,
    class_date: Optional[date] = None,
    current=Depends(auth.require_role("teacher")),
    db: Session = Depends(get_db)
):
    q = db.query(models.AttendanceRecord).filter(models.AttendanceRecord.subject_id == subject_id)
    if class_date:
        q = q.filter(models.AttendanceRecord.class_date == class_date)
    recs = q.all()
    return [{"student_id": r.student_id, "class_date": str(r.class_date), "status": r.status} for r in recs]


@app.get("/api/teacher/export/{subject_id}")
def export_attendance(
    subject_id: int,
    current=Depends(auth.require_role("teacher")),
    db: Session = Depends(get_db)
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule

    teacher = current["user"]
    subj = db.query(models.Subject).filter(
        models.Subject.id == subject_id,
        models.Subject.teacher_id == teacher.id
    ).first()
    if not subj:
        raise HTTPException(404, "Subject not found")

    branch = db.query(models.Branch).filter(models.Branch.id == subj.branch_id).first()
    batch = db.query(models.Batch).filter(models.Batch.id == subj.batch_id).first()

    dates = db.query(models.AttendanceRecord.class_date).filter(
        models.AttendanceRecord.subject_id == subject_id
    ).distinct().order_by(models.AttendanceRecord.class_date).all()
    dates = [d[0] for d in dates]

    enrollments = db.query(models.Enrollment).filter(models.Enrollment.subject_id == subject_id).all()
    student_ids = [e.student_id for e in enrollments]
    students = db.query(models.Student).filter(
        models.Student.id.in_(student_ids)
    ).all()

    # Group students by type → back_year, lateral_entry, regular (in that order as per sample)
    def sort_key(s):
        reg = s.reg_no or ""
        return reg
    regular = sorted([s for s in students if s.student_type == "regular"], key=sort_key)
    lateral = sorted([s for s in students if s.student_type == "lateral_entry"], key=sort_key)
    back    = sorted([s for s in students if s.student_type == "back_year"], key=sort_key)

    # Group back-year by batch prefix (21 Batch, 22 Batch ...)
    back_groups = {}
    for s in back:
        key = s.reg_no[:2] if s.reg_no else "??"
        back_groups.setdefault(key, []).append(s)

    records = db.query(models.AttendanceRecord).filter(
        models.AttendanceRecord.subject_id == subject_id
    ).all()
    rec_map = {(r.student_id, r.class_date): r.status for r in records}

    total_days = len(dates)
    ncols = 4 + total_days + 4   # S.No, Reg No, Name, Branch + dates + Present, Absent, Total, %
    branch_code = branch.code if branch else "-"

    wb = Workbook()
    ws = wb.active
    ws.title = f"{subj.code}"[:31]

    # ===== STYLES =====
    title_fill   = PatternFill("solid", start_color="4F46E5")   # Indigo
    sub_fill     = PatternFill("solid", start_color="6366F1")   # Indigo light
    hdr_fill     = PatternFill("solid", start_color="1E40AF")   # Deep blue
    group_fill   = PatternFill("solid", start_color="FEF3C7")   # Yellow-50
    present_fill = PatternFill("solid", start_color="D1FAE5")
    absent_fill  = PatternFill("solid", start_color="FEE2E2")
    footer_fill  = PatternFill("solid", start_color="F1F5F9")
    thin = Side(border_style="thin", color="94A3B8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ===== Row 1 — Subject title =====
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=f"{subj.name} - Student Attendance Report")
    c.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    c.fill = title_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ===== Row 2 — Subject details =====
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    from_d = dates[0].strftime("%d-%b") if dates else "-"
    to_d = dates[-1].strftime("%d-%b %Y") if dates else "-"
    sub_info = (
        f"{_ordinal(subj.semester)} Semester, {branch_code} ({batch.name if batch else '-'})  "
        f"•  Duration: {from_d} – {to_d}  •  Total Classes: {total_days}"
    )
    c = ws.cell(row=2, column=1, value=sub_info)
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = sub_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # Row 3 is empty gap

    # ===== Row 4 — Column headers =====
    headers = ["S.No", "Registration No", "Name", "Branch"]
    headers += [d.strftime("%d-%b") for d in dates]
    headers += ["Present", "Absent", "Total", "Attendance %"]

    for i, h in enumerate(headers, 1):
        cc = ws.cell(row=4, column=i, value=h)
        cc.font = Font(bold=True, color="FFFFFF", size=10)
        cc.fill = hdr_fill
        cc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cc.border = border
    ws.row_dimensions[4].height = 34

    # ===== DATA ROWS with grouping =====
    current_row = 5
    s_no = 0

    def write_group_header(label):
        nonlocal current_row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=ncols)
        gc = ws.cell(row=current_row, column=1, value=f"  {label}")
        gc.font = Font(bold=True, size=10, italic=True, color="92400E")
        gc.fill = group_fill
        gc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    def write_student(s):
        nonlocal current_row, s_no
        s_no += 1
        r = current_row
        # Static columns
        ws.cell(row=r, column=1, value=s_no).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2, value=s.reg_no).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=2).border = border
        ws.cell(row=r, column=3, value=s.name).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=r, column=3).border = border
        ws.cell(row=r, column=4, value=branch_code).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r, column=4).border = border
        for ci in range(1, 5):
            ws.cell(row=r, column=ci).font = Font(size=10)

        # Date columns — P or A
        for di, d in enumerate(dates):
            status = rec_map.get((s.id, d), "A")
            cc = ws.cell(row=r, column=5 + di, value=status)
            cc.alignment = Alignment(horizontal="center", vertical="center")
            cc.fill = present_fill if status == "P" else absent_fill
            cc.font = Font(size=10, bold=True, color="059669" if status == "P" else "DC2626")
            cc.border = border

        # Formula columns — Present, Absent, Total, %
        first_date_col = 5
        last_date_col = 4 + total_days
        first_letter = get_column_letter(first_date_col)
        last_letter = get_column_letter(last_date_col)

        present_col = last_date_col + 1
        absent_col  = last_date_col + 2
        total_col   = last_date_col + 3
        pct_col     = last_date_col + 4

        cp = ws.cell(row=r, column=present_col,
                     value=f'=COUNTIF({first_letter}{r}:{last_letter}{r},"P")')
        cp.alignment = Alignment(horizontal="center", vertical="center")
        cp.border = border
        cp.font = Font(bold=True, size=10, color="059669")

        ca = ws.cell(row=r, column=absent_col,
                     value=f'=COUNTIF({first_letter}{r}:{last_letter}{r},"A")')
        ca.alignment = Alignment(horizontal="center", vertical="center")
        ca.border = border
        ca.font = Font(bold=True, size=10, color="DC2626")

        ct = ws.cell(row=r, column=total_col, value=total_days)
        ct.alignment = Alignment(horizontal="center", vertical="center")
        ct.border = border
        ct.font = Font(size=10)

        present_letter = get_column_letter(present_col)
        total_letter   = get_column_letter(total_col)
        cpct = ws.cell(row=r, column=pct_col,
                       value=f'={present_letter}{r}/{total_letter}{r}')
        cpct.alignment = Alignment(horizontal="center", vertical="center")
        cpct.border = border
        # Per-bucket colour: green ≥ 75%, amber 60–<75%, red < 60%.
        # Excel formulas haven't been evaluated at write time, so we use
        # workbook-level conditional formatting (added once after the loop)
        # for the colours and just bold the value here.
        cpct.font = Font(bold=True, size=10)
        cpct.number_format = "0.0%"

        current_row += 1

    # Back-year groups first (oldest first)
    for yr in sorted(back_groups.keys()):
        write_group_header(f"{yr} Batch (Back Year)")
        for s in back_groups[yr]:
            write_student(s)

    # Lateral Entry next
    if lateral:
        # Detect LE batch (mostly 24)
        le_yr = lateral[0].reg_no[:2] if lateral[0].reg_no else "24"
        write_group_header(f"Lateral Entry ({le_yr} Batch)")
        for s in lateral:
            write_student(s)

    # Regular students last
    if regular:
        reg_yr = regular[0].reg_no[:2] if regular[0].reg_no else "23"
        write_group_header(f"Regular Students ({reg_yr} Batch)")
        for s in regular:
            write_student(s)

    # Footer
    footer_row = current_row + 1
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=ncols)
    fc = ws.cell(row=footer_row, column=1,
                 value=f"Generated: {now_ist().strftime('%d %b %Y, %I:%M %p')}  •  "
                       f"Teacher: {teacher.name}  •  Class Attendance v1.2.0  •  RRSDCE Begusarai")
    fc.font = Font(size=9, italic=True, color="64748B")
    fc.fill = footer_fill
    fc.alignment = Alignment(horizontal="center", vertical="center")

    # ===== Conditional formatting for the % column =====
    # Bucket colours requested:
    #   ≥ 75%      → safe (green)
    #   60–<75%    → warning (amber)
    #   < 60%      → critical (red)
    # Cells contain a fraction (0–1) because of the "0.0%" format, so
    # thresholds are 0.75 and 0.60.
    if current_row > 5:
        pct_col_letter = get_column_letter(4 + total_days + 4)
        pct_range = f"{pct_col_letter}5:{pct_col_letter}{current_row - 1}"
        green_fill  = PatternFill("solid", start_color="D1FAE5")
        green_font  = Font(bold=True, color="047857", size=10)
        amber_fill  = PatternFill("solid", start_color="FEF3C7")
        amber_font  = Font(bold=True, color="B45309", size=10)
        red_fill    = PatternFill("solid", start_color="FEE2E2")
        red_font    = Font(bold=True, color="B91C1C", size=10)
        ws.conditional_formatting.add(
            pct_range,
            CellIsRule(operator="greaterThanOrEqual", formula=["0.75"],
                       fill=green_fill, font=green_font, stopIfTrue=True),
        )
        ws.conditional_formatting.add(
            pct_range,
            CellIsRule(operator="greaterThanOrEqual", formula=["0.60"],
                       fill=amber_fill, font=amber_font, stopIfTrue=True),
        )
        ws.conditional_formatting.add(
            pct_range,
            CellIsRule(operator="lessThan", formula=["0.60"],
                       fill=red_fill, font=red_font, stopIfTrue=True),
        )

    # Column widths
    ws.column_dimensions['A'].width = 6    # S.No
    ws.column_dimensions['B'].width = 16   # Reg No
    ws.column_dimensions['C'].width = 26   # Name
    ws.column_dimensions['D'].width = 9    # Branch
    for i in range(total_days):
        ws.column_dimensions[get_column_letter(5 + i)].width = 8
    ws.column_dimensions[get_column_letter(4 + total_days + 1)].width = 9   # Present
    ws.column_dimensions[get_column_letter(4 + total_days + 2)].width = 9   # Absent
    ws.column_dimensions[get_column_letter(4 + total_days + 3)].width = 9   # Total
    ws.column_dimensions[get_column_letter(4 + total_days + 4)].width = 12  # %

    # Freeze header + first 4 columns
    ws.freeze_panes = "E5"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = subj.name.replace(' ', '_').replace('/', '_')
    fname = f"{safe_name}({subj.code})_{batch.name if batch else 'BATCH'}_Sem{subj.semester}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )


def _ordinal(n):
    """Convert int to ordinal string like 5 → 5th"""
    if n is None:
        return ""
    if 10 <= (n % 100) <= 20:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suffix}"


@app.get("/api/teacher/export-pdf/{subject_id}")
def export_attendance_pdf(
    subject_id: int,
    current=Depends(auth.require_role("teacher")),
    db: Session = Depends(get_db)
):
    from reportlab.lib.pagesizes import landscape, A3
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import mm

    teacher = current["user"]
    subj = db.query(models.Subject).filter(
        models.Subject.id == subject_id,
        models.Subject.teacher_id == teacher.id
    ).first()
    if not subj:
        raise HTTPException(404, "Subject not found")

    branch = db.query(models.Branch).filter(models.Branch.id == subj.branch_id).first()
    batch = db.query(models.Batch).filter(models.Batch.id == subj.batch_id).first()

    dates = db.query(models.AttendanceRecord.class_date).filter(
        models.AttendanceRecord.subject_id == subject_id
    ).distinct().order_by(models.AttendanceRecord.class_date).all()
    dates = [d[0] for d in dates]

    enrollments = db.query(models.Enrollment).filter(models.Enrollment.subject_id == subject_id).all()
    student_ids = [e.student_id for e in enrollments]
    students = db.query(models.Student).filter(models.Student.id.in_(student_ids)).all()

    def sort_key(s): return s.reg_no or ""
    regular = sorted([s for s in students if s.student_type == "regular"], key=sort_key)
    lateral = sorted([s for s in students if s.student_type == "lateral_entry"], key=sort_key)
    back    = sorted([s for s in students if s.student_type == "back_year"], key=sort_key)
    back_groups = {}
    for s in back:
        key = s.reg_no[:2] if s.reg_no else "??"
        back_groups.setdefault(key, []).append(s)

    records = db.query(models.AttendanceRecord).filter(
        models.AttendanceRecord.subject_id == subject_id
    ).all()
    rec_map = {(r.student_id, r.class_date): r.status for r in records}

    branch_code = branch.code if branch else "-"
    total_days = len(dates)

    headers = ["S.No", "Reg No", "Name", "Br"] + [d.strftime("%d-%b") for d in dates] + ["P", "A", "T", "%"]
    data = [headers]

    s_no = 0
    group_rows = []
    pct_color_rows = []  # (row_index_in_data, color_hex) for % column shading

    def add_group(label):
        group_rows.append(len(data))
        data.append([label] + [""] * (len(headers) - 1))

    def add_student(s):
        nonlocal s_no
        s_no += 1
        present = sum(1 for d in dates if rec_map.get((s.id, d), "A") == "P")
        absent = total_days - present
        pct = (present / total_days * 100) if total_days else 0
        row = [s_no, s.reg_no, s.name, branch_code]
        row += [rec_map.get((s.id, d), "A") for d in dates]
        row += [present, absent, total_days, f"{pct:.1f}%"]
        data.append(row)
        # Bucket colour: ≥75 green, 60–<75 amber, <60 red
        if pct >= 75:
            pct_color_rows.append((len(data) - 1, "#047857", "#D1FAE5"))
        elif pct >= 60:
            pct_color_rows.append((len(data) - 1, "#B45309", "#FEF3C7"))
        else:
            pct_color_rows.append((len(data) - 1, "#B91C1C", "#FEE2E2"))

    for yr in sorted(back_groups.keys()):
        add_group(f"  {yr} Batch (Back Year)")
        for s in back_groups[yr]: add_student(s)
    if lateral:
        le_yr = lateral[0].reg_no[:2] if lateral[0].reg_no else "24"
        add_group(f"  Lateral Entry ({le_yr} Batch)")
        for s in lateral: add_student(s)
    if regular:
        reg_yr = regular[0].reg_no[:2] if regular[0].reg_no else "23"
        add_group(f"  Regular Students ({reg_yr} Batch)")
        for s in regular: add_student(s)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A3),
        leftMargin=10*mm, rightMargin=10*mm, topMargin=10*mm, bottomMargin=10*mm
    )
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('t', parent=styles['Heading1'], fontSize=15,
                                 textColor=colors.HexColor('#FFFFFF'), alignment=1,
                                 backColor=colors.HexColor('#4F46E5'), borderPadding=8)
    sub_style = ParagraphStyle('s', parent=styles['Normal'], fontSize=10,
                               textColor=colors.HexColor('#FFFFFF'), alignment=1,
                               backColor=colors.HexColor('#6366F1'), borderPadding=6)

    from_d = dates[0].strftime("%d-%b") if dates else "-"
    to_d = dates[-1].strftime("%d-%b %Y") if dates else "-"
    elements.append(Paragraph(f"<b>{subj.name} - Student Attendance Report</b>", title_style))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph(
        f"{_ordinal(subj.semester)} Semester, {branch_code} ({batch.name if batch else '-'})  "
        f"&bull;  Duration: {from_d} – {to_d}  &bull;  Total Classes: {total_days}",
        sub_style
    ))
    elements.append(Spacer(1, 8))

    col_widths = [10*mm, 22*mm, 45*mm, 10*mm] + [8*mm] * total_days + [9*mm, 9*mm, 9*mm, 12*mm]
    tbl = Table(data, colWidths=col_widths, repeatRows=1)

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, 0), 8),
        ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN',     (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE',   (0, 1), (-1, -1), 7.5),
        ('GRID',       (0, 0), (-1, -1), 0.3, colors.HexColor('#94A3B8')),
        ('ALIGN',      (0, 1), (-1, -1), 'CENTER'),
        ('ALIGN',      (2, 1), (2, -1), 'LEFT'),
    ])
    for gr in group_rows:
        style.add('BACKGROUND', (0, gr), (-1, gr), colors.HexColor('#FEF3C7'))
        style.add('TEXTCOLOR',  (0, gr), (-1, gr), colors.HexColor('#92400E'))
        style.add('FONTNAME',   (0, gr), (-1, gr), 'Helvetica-Bold')
        style.add('SPAN',       (0, gr), (-1, gr))
        style.add('ALIGN',      (0, gr), (-1, gr), 'LEFT')

    for r in range(1, len(data)):
        if r in group_rows: continue
        for ci in range(4, 4 + total_days):
            val = data[r][ci]
            if val == "P":
                style.add('BACKGROUND', (ci, r), (ci, r), colors.HexColor('#D1FAE5'))
                style.add('TEXTCOLOR',  (ci, r), (ci, r), colors.HexColor('#059669'))
            elif val == "A":
                style.add('BACKGROUND', (ci, r), (ci, r), colors.HexColor('#FEE2E2'))
                style.add('TEXTCOLOR',  (ci, r), (ci, r), colors.HexColor('#DC2626'))
        style.add('FONTNAME', (4, r), (4 + total_days - 1, r), 'Helvetica-Bold')

    # Per-row colouring for the % column (last column)
    pct_col_idx = len(headers) - 1
    for r, txt, bg in pct_color_rows:
        style.add('BACKGROUND', (pct_col_idx, r), (pct_col_idx, r), colors.HexColor(bg))
        style.add('TEXTCOLOR',  (pct_col_idx, r), (pct_col_idx, r), colors.HexColor(txt))
        style.add('FONTNAME',   (pct_col_idx, r), (pct_col_idx, r), 'Helvetica-Bold')

    tbl.setStyle(style)
    elements.append(tbl)

    elements.append(Spacer(1, 8))
    footer_style = ParagraphStyle('f', parent=styles['Normal'], fontSize=8,
                                  textColor=colors.HexColor('#64748B'), alignment=1)
    elements.append(Paragraph(
        f"Generated: {now_ist().strftime('%d %b %Y, %I:%M %p')}  &bull;  "
        f"Teacher: {teacher.name}  &bull;  Class Attendance v1.2.0  &bull;  RRSDCE Begusarai",
        footer_style
    ))

    doc.build(elements)
    buf.seek(0)
    safe_name = subj.name.replace(' ', '_').replace('/', '_')
    fname = f"{safe_name}({subj.code})_{batch.name if batch else 'BATCH'}_Sem{subj.semester}.pdf"
    return Response(
        content=buf.read(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )


# ============ STUDENT ============
@app.get("/api/student/attendance")
def my_attendance(current=Depends(auth.require_role("student")), db: Session = Depends(get_db)):
    """Student dashboard. Optimised to ~5 DB queries flat — was doing 4 queries
    per enrolled subject (10+ subjects → 40+ round-trips, painfully slow on
    Render cold starts)."""
    student = current["user"]
    batch = db.query(models.Batch).filter(models.Batch.id == student.batch_id).first()
    branch = db.query(models.Branch).filter(models.Branch.id == student.branch_id).first()

    # 1. Fetch enrollments + subjects in one JOIN
    subjects = db.query(models.Subject).join(
        models.Enrollment, models.Enrollment.subject_id == models.Subject.id
    ).filter(models.Enrollment.student_id == student.id).all()
    if not subjects:
        return {
            "student": {
                "reg_no": student.reg_no, "name": student.name,
                "semester": student.current_semester,
                "student_type": student.student_type,
                "is_verified": student.is_verified,
                "batch": batch.name if batch else None,
                "branch": branch.name if branch else None,
                "branch_code": branch.code if branch else None,
            },
            "subjects": [],
        }

    subject_ids = [s.id for s in subjects]

    # 2. All teacher names for the subjects (one query)
    teacher_ids = {s.teacher_id for s in subjects if s.teacher_id}
    teacher_map = {
        t.id: t.name for t in
        db.query(models.Teacher).filter(models.Teacher.id.in_(teacher_ids)).all()
    } if teacher_ids else {}

    # 3. Total class days per subject — one GROUP BY query
    total_rows = db.query(
        models.AttendanceRecord.subject_id,
        func.count(func.distinct(models.AttendanceRecord.class_date)),
    ).filter(
        models.AttendanceRecord.subject_id.in_(subject_ids)
    ).group_by(models.AttendanceRecord.subject_id).all()
    total_map = {sid: count for sid, count in total_rows}

    # 4. Present count per subject for THIS student — one GROUP BY query
    present_rows = db.query(
        models.AttendanceRecord.subject_id,
        func.count(models.AttendanceRecord.id),
    ).filter(
        models.AttendanceRecord.subject_id.in_(subject_ids),
        models.AttendanceRecord.student_id == student.id,
        models.AttendanceRecord.status == "P",
    ).group_by(models.AttendanceRecord.subject_id).all()
    present_map = {sid: count for sid, count in present_rows}

    result = []
    for subj in subjects:
        actual_classes = total_map.get(subj.id, 0)
        present = present_map.get(subj.id, 0)
        pct = (present / actual_classes * 100) if actual_classes else 0
        classes_needed = max(0, int((0.75 * actual_classes - present) / 0.25)) if actual_classes else 0
        result.append({
            "subject_id": subj.id,
            "subject_name": subj.name,
            "subject_code": subj.code,
            "semester": subj.semester,
            "credits": subj.credits,
            "teacher_name": teacher_map.get(subj.teacher_id),
            "total_classes": actual_classes,
            "present": present,
            "absent": actual_classes - present,
            "percentage": round(pct, 2),
            "classes_needed_for_75": classes_needed,
        })
    return {
        "student": {
            "reg_no": student.reg_no,
            "name": student.name,
            "semester": student.current_semester,
            "student_type": student.student_type,
            "is_verified": student.is_verified,
            "batch": batch.name if batch else None,
            "branch": branch.name if branch else None,
            "branch_code": branch.code if branch else None,
        },
        "subjects": result
    }


@app.get("/api/student/attendance/{subject_id}")
def my_subject_detail(
    subject_id: int,
    current=Depends(auth.require_role("student")),
    db: Session = Depends(get_db)
):
    student = current["user"]
    subj = db.query(models.Subject).filter(models.Subject.id == subject_id).first()
    if not subj:
        raise HTTPException(404, "Subject not found")

    records = db.query(models.AttendanceRecord).filter(
        models.AttendanceRecord.subject_id == subject_id,
        models.AttendanceRecord.student_id == student.id
    ).order_by(models.AttendanceRecord.class_date).all()

    total = db.query(func.count(func.distinct(models.AttendanceRecord.class_date))).filter(
        models.AttendanceRecord.subject_id == subject_id
    ).scalar() or 0
    present = sum(1 for r in records if r.status == "P")
    absent = sum(1 for r in records if r.status == "A")
    pct = (present / total * 100) if total else 0

    # Analysis
    present_dates = [str(r.class_date) for r in records if r.status == "P"]
    absent_dates = [str(r.class_date) for r in records if r.status == "A"]

    # streak analysis
    streak = 0
    max_absent_streak = 0
    cur_abs = 0
    for r in records:
        if r.status == "A":
            cur_abs += 1
            max_absent_streak = max(max_absent_streak, cur_abs)
        else:
            cur_abs = 0

    # Teacher info for subject heading
    teacher = db.query(models.Teacher).filter(models.Teacher.id == subj.teacher_id).first()

    return {
        "subject": {
            "id": subj.id,
            "code": subj.code,
            "name": subj.name,
            "semester": subj.semester,
            "credits": subj.credits,
            "session": subj.session,
            "teacher_name": teacher.name if teacher else None,
        },
        "summary": {
            "total_classes": total,
            "present": present,
            "absent": absent,
            "percentage": round(pct, 2),
        },
        "analysis": {
            "max_absent_streak": max_absent_streak,
            "classes_needed_for_75": max(0, int((0.75 * total - present) / 0.25)) if total else 0,
            # Buckets must match Excel/PDF + teacher dashboard:
            #   Safe ≥ 75   |   Warning 60-<75   |   Critical < 60
            "status": "Safe" if pct >= 75 else ("Warning" if pct >= 60 else "Critical"),
        },
        "present_dates": present_dates,
        "absent_dates": absent_dates,
        "all_records": [
            {
                "date": str(r.class_date),
                "status": r.status,
                "marked_at": r.marked_at.isoformat() if r.marked_at else None,
            } for r in records
        ],
    }


# ============ BRANCHES CRUD (Super Admin) ============
@app.post("/api/superadmin/branches", response_model=schemas.BranchOut)
def create_branch(
    data: schemas.BranchCreate,
    current=Depends(auth.require_role("superadmin")),
    db: Session = Depends(get_db)
):
    existing = db.query(models.Branch).filter(models.Branch.code == data.code).first()
    if existing:
        raise HTTPException(400, "Branch code already exists")
    b = models.Branch(name=data.name, code=data.code)
    db.add(b)
    db.commit()
    db.refresh(b)
    return b


@app.patch("/api/superadmin/branches/{branch_id}", response_model=schemas.BranchOut)
def update_branch(
    branch_id: int,
    data: schemas.BranchUpdate,
    current=Depends(auth.require_role("superadmin")),
    db: Session = Depends(get_db)
):
    b = db.query(models.Branch).filter(models.Branch.id == branch_id).first()
    if not b:
        raise HTTPException(404, "Branch not found")
    if data.name is not None: b.name = data.name
    if data.code is not None: b.code = data.code
    db.commit()
    db.refresh(b)
    return b


@app.delete("/api/superadmin/branches/{branch_id}")
def delete_branch(
    branch_id: int,
    current=Depends(auth.require_role("superadmin")),
    db: Session = Depends(get_db)
):
    b = db.query(models.Branch).filter(models.Branch.id == branch_id).first()
    if not b:
        raise HTTPException(404, "Branch not found")
    # Block delete if branch has students, teachers, admins
    has_data = (
        db.query(models.Student).filter(models.Student.branch_id == branch_id).first()
        or db.query(models.Teacher).filter(models.Teacher.branch_id == branch_id).first()
        or db.query(models.Admin).filter(models.Admin.branch_id == branch_id).first()
    )
    if has_data:
        raise HTTPException(400, "Cannot delete: branch has admins/teachers/students. Remove them first.")
    db.delete(b)
    db.commit()
    return {"success": True}


# ============ BATCH UPDATE (Super Admin) ============
@app.patch("/api/superadmin/batches/{batch_id}", response_model=schemas.BatchOut)
def update_batch(
    batch_id: int,
    data: schemas.BatchUpdate,
    current=Depends(auth.require_role("superadmin")),
    db: Session = Depends(get_db)
):
    b = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
    if not b:
        raise HTTPException(404, "Batch not found")
    if data.current_semester is not None: b.current_semester = data.current_semester
    if data.is_active is not None: b.is_active = data.is_active
    db.commit()
    db.refresh(b)
    return b


@app.delete("/api/superadmin/batches/{batch_id}")
def delete_batch(
    batch_id: int,
    current=Depends(auth.require_role("superadmin")),
    db: Session = Depends(get_db)
):
    b = db.query(models.Batch).filter(models.Batch.id == batch_id).first()
    if not b:
        raise HTTPException(404, "Batch not found")
    has_students = db.query(models.Student).filter(models.Student.batch_id == batch_id).first()
    if has_students:
        raise HTTPException(400, "Cannot delete: batch has enrolled students")
    db.delete(b)
    db.commit()
    return {"success": True}


# ============ DRILL-DOWN VIEWS (Super Admin clickable cards) ============
# These endpoints power the "tap a card → see all rows" listings on the
# super-admin dashboard. They MUST be fast even for thousands of rows
# (Render free-tier + Supabase pooler can chug on N+1 queries).
# Strategy: pre-fetch the lookup tables (branches, batches) into dicts,
# then build the response in pure-Python — exactly 2-4 queries total
# regardless of row count.

@app.get("/api/superadmin/all-teachers")
def all_teachers(current=Depends(auth.require_role("superadmin")), db: Session = Depends(get_db)):
    branch_map = {b.id: b for b in db.query(models.Branch).all()}
    # Pre-build "teacher_id → list of branch codes" via TeacherBranch so we
    # can show "CSE, CSE-DS" for cross-branch teachers in a single column.
    tb_rows = db.query(models.TeacherBranch).all()
    teacher_branches = {}
    for tb in tb_rows:
        teacher_branches.setdefault(tb.teacher_id, []).append(tb.branch_id)

    teachers = db.query(models.Teacher).order_by(models.Teacher.name).all()
    result = []
    for t in teachers:
        b_ids = teacher_branches.get(t.id) or ([t.branch_id] if t.branch_id else [])
        codes = [branch_map[bid].code for bid in b_ids if bid in branch_map]
        # Primary branch (the home/first one) — used for the colored badge in UI
        primary = branch_map.get(t.branch_id)
        result.append({
            "id": t.id, "name": t.name, "email": t.email,
            "phone": t.phone, "is_verified": t.is_verified,
            "branch_code": primary.code if primary else (codes[0] if codes else None),
            "branch_name": primary.name if primary else None,
            "branch_codes": codes,  # full list for cross-branch display
        })
    return result


@app.get("/api/superadmin/all-students")
def all_students(current=Depends(auth.require_role("superadmin")), db: Session = Depends(get_db)):
    branch_map = {b.id: b.code for b in db.query(models.Branch).all()}
    batch_map = {b.id: b.name for b in db.query(models.Batch).all()}
    students = db.query(models.Student).order_by(models.Student.reg_no).all()
    return [{
        "id": s.id, "reg_no": s.reg_no, "name": s.name,
        "branch_code": branch_map.get(s.branch_id),
        "batch": batch_map.get(s.batch_id),
        "semester": s.current_semester,
        "student_type": s.student_type,
        "is_verified": s.is_verified,
    } for s in students]


@app.get("/api/superadmin/all-subjects")
def all_subjects(current=Depends(auth.require_role("superadmin")), db: Session = Depends(get_db)):
    branch_map = {b.id: b.code for b in db.query(models.Branch).all()}
    batch_map = {b.id: b.name for b in db.query(models.Batch).all()}
    teacher_map = {t.id: t.name for t in db.query(models.Teacher).all()}
    subjects = db.query(models.Subject).order_by(models.Subject.code).all()
    return [{
        "id": s.id, "code": s.code, "name": s.name,
        "branch_code": branch_map.get(s.branch_id),
        "batch": batch_map.get(s.batch_id),
        "semester": s.semester,
        "credits": s.credits,
        "session": s.session,
        "teacher_name": teacher_map.get(s.teacher_id),
    } for s in subjects]


# ============ TEACHER ATTENDANCE DASHBOARD ============
@app.get("/api/teacher/dashboard/{subject_id}")
def teacher_dashboard(
    subject_id: int,
    current=Depends(auth.require_role("teacher")),
    db: Session = Depends(get_db)
):
    """Teacher dashboard: per-student attendance % + recent classes + stats.

    Optimised to 4 DB queries flat (subject + dates + enrollments+students JOIN +
    all attendance records aggregated in Python). Previous version did one query
    per enrolled student → 50+ round-trips per dashboard load on Render free tier
    + Supabase pooler, which made the page take 10-15s on warm hits."""
    teacher = current["user"]
    subject = db.query(models.Subject).filter(
        models.Subject.id == subject_id,
        models.Subject.teacher_id == teacher.id
    ).first()
    if not subject:
        raise HTTPException(404, "Subject not found")

    # 1. All distinct class dates for this subject
    all_dates = db.query(models.AttendanceRecord.class_date).filter(
        models.AttendanceRecord.subject_id == subject_id
    ).distinct().order_by(models.AttendanceRecord.class_date.desc()).all()
    class_dates = [str(d[0]) for d in all_dates]
    total_classes = len(class_dates)

    # 2. Enrolled students — single JOIN, no N+1
    enrolled_students = db.query(models.Student).join(
        models.Enrollment, models.Enrollment.student_id == models.Student.id
    ).filter(models.Enrollment.subject_id == subject_id).all()

    # 3. ALL attendance records for this subject in one shot, aggregated in Python
    records = db.query(
        models.AttendanceRecord.student_id,
        models.AttendanceRecord.status,
    ).filter(models.AttendanceRecord.subject_id == subject_id).all()

    # student_id → {present, total}
    counts = {}
    for sid, status in records:
        c = counts.setdefault(sid, {"present": 0, "total": 0})
        c["total"] += 1
        if status == "P":
            c["present"] += 1

    students_stats = []
    for s in enrolled_students:
        c = counts.get(s.id, {"present": 0, "total": 0})
        present, total = c["present"], c["total"]
        pct = (present / total * 100) if total else 0
        students_stats.append({
            "student_id": s.id,
            "reg_no": s.reg_no,
            "name": s.name,
            "student_type": s.student_type,
            "present": present,
            "total": total,
            "percentage": round(pct, 1),
            "is_verified": s.is_verified,
        })
    students_stats.sort(key=lambda x: -x['percentage'])

    return {
        "subject": {
            "id": subject.id,
            "code": subject.code,
            "name": subject.name,
            "semester": subject.semester,
            "credits": subject.credits,
            "session": subject.session,
        },
        "total_classes": total_classes,
        "recent_dates": class_dates[:10],
        "all_dates": class_dates,
        "students": students_stats,
    }


# ============ TEACHER: BACK-DATE EDIT ATTENDANCE ============
@app.post("/api/teacher/attendance/edit")
def edit_attendance(
    data: schemas.AttendanceEdit,
    current=Depends(auth.require_role("teacher")),
    db: Session = Depends(get_db)
):
    """Allow teacher to modify past attendance for a specific date."""
    teacher = current["user"]
    subject = db.query(models.Subject).filter(
        models.Subject.id == data.subject_id,
        models.Subject.teacher_id == teacher.id
    ).first()
    if not subject:
        raise HTTPException(404, "Subject not found or not assigned to you")

    # Delete existing records for that date
    db.query(models.AttendanceRecord).filter(
        models.AttendanceRecord.subject_id == data.subject_id,
        models.AttendanceRecord.class_date == data.class_date
    ).delete()

    # All-absent edit removes the day completely (same rule as /mark).
    has_present = any(m.get("status") == "P" for m in data.marks)
    if not has_present:
        db.commit()
        return {
            "success": True,
            "message": "Attendance for this date removed",
            "date": str(data.class_date),
            "deleted_day": True,
        }

    # Re-insert with updated data
    for m in data.marks:
        rec = models.AttendanceRecord(
            student_id=m['student_id'],
            subject_id=data.subject_id,
            class_date=data.class_date,
            status=m['status'],
            marked_by_teacher_id=teacher.id,
            marked_at=datetime.utcnow(),
        )
        db.add(rec)
    db.commit()
    return {"success": True, "message": "Attendance updated", "date": str(data.class_date)}


# ============ TEACHER: UPDATE OWN PROFILE (image etc) ============
@app.get("/api/teacher/me", response_model=schemas.TeacherOut)
def teacher_me(
    current=Depends(auth.require_role("teacher")),
    db: Session = Depends(get_db)
):
    return current["user"]


@app.patch("/api/teacher/profile", response_model=schemas.TeacherOut)
def update_teacher_self(
    data: schemas.TeacherUpdate,
    current=Depends(auth.require_role("teacher")),
    db: Session = Depends(get_db)
):
    teacher = current["user"]
    if data.name is not None: teacher.name = data.name
    if data.phone is not None: teacher.phone = data.phone
    if data.profile_image is not None: teacher.profile_image = data.profile_image
    db.commit()
    db.refresh(teacher)
    return teacher

@app.get("/api/teacher/attendance/{subject_id}/dates")
def get_marked_dates(
    subject_id: int,
    current=Depends(auth.require_role("teacher")),
    db: Session = Depends(get_db),
):
    teacher = current["user"]
    subj = db.query(models.Subject).filter(
        models.Subject.id == subject_id,
        models.Subject.teacher_id == teacher.id,
    ).first()
    if not subj:
        raise HTTPException(404, "Subject not assigned to you")
    rows = db.query(models.AttendanceRecord.class_date).filter(
        models.AttendanceRecord.subject_id == subject_id
    ).distinct().all()
    return {"dates": [str(r[0]) for r in rows]}